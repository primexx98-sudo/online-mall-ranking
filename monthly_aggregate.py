from pathlib import Path
from datetime import date
from collections import defaultdict
from crawlers.classifier import classify, ALL_CATEGORIES, PLATFORMS as PLATFORM_NAMES
import pandas as pd


PLATFORMS   = ["카카오선물하기", "다이소몰", "올리브영"]
COLUMNS_OUT = [
    "순위(월평균)", "카테고리", "상품명", "브랜드",
    "평균가격", "전월가격", "가격변동",
    "평균순위", "전월순위", "순위변동",
    "등장횟수",
]


def price_to_int(price_str: str) -> float:
    try:
        digits = "".join(c for c in str(price_str) if c.isdigit())
        return float(digits) if digits else 0.0
    except Exception:
        return 0.0


def parse_saved_price(price_str):
    """월별취합 시트에 저장된 '12,345원' 형식을 정수로 역변환 (전월 비교용)."""
    if not price_str or price_str == "-":
        return None
    digits = "".join(c for c in str(price_str) if c.isdigit())
    return int(digits) if digits else None


def load_previous_month(target_year: int, target_month: int) -> dict:
    """전월 월별취합.xlsx의 모든 시트를 dict로 로드 (없으면 빈 dict)."""
    if target_month == 1:
        prev_year, prev_month = target_year - 1, 12
    else:
        prev_year, prev_month = target_year, target_month - 1
    prev_path = Path("data/monthly") / f"{prev_year}-{prev_month:02d}_월별취합.xlsx"
    if not prev_path.exists():
        return {}
    try:
        return pd.read_excel(prev_path, sheet_name=None)
    except Exception:
        return {}


def rank_change_label(prev_rank, curr_rank) -> str:
    if prev_rank is None:
        return "🆕"
    diff = int(prev_rank) - int(curr_rank)
    if diff > 0:
        return f"▲{diff}"
    if diff < 0:
        return f"▼{-diff}"
    return "-"


def price_change_label(prev_price, curr_price) -> str:
    if prev_price is None or curr_price is None or curr_price <= 0:
        return "-"
    diff = int(curr_price) - int(prev_price)
    if diff > 0:
        return f"▲{diff:,}원"
    if diff < 0:
        return f"▼{abs(diff):,}원"
    return "-"


def aggregate_platform(daily_files: list, sheet: str, total_days: int, prev_df: pd.DataFrame = None) -> pd.DataFrame:
    frames = []
    for f in daily_files:
        try:
            df = pd.read_excel(f, sheet_name=sheet)
            df["_date"] = f.stem
            frames.append(df)
        except Exception:
            pass

    if not frames:
        return pd.DataFrame(columns=COLUMNS_OUT)

    all_data = pd.concat(frames, ignore_index=True)
    all_data["가격_숫자"] = all_data["가격"].apply(price_to_int)
    # 일별 점수: 1위=5점, 2위=4.5점 ... 10위=0.5점
    all_data["일별점수"] = all_data["순위"].apply(lambda r: (11 - r) * 0.5)

    grouped = all_data.groupby(["상품명", "브랜드"]).agg(
        평균순위=("순위", "mean"),
        등장횟수=("순위", "count"),
        평균가격_raw=("가격_숫자", "mean"),
        점수합=("일별점수", "sum"),
    ).reset_index()

    # 등장횟수 3회 미만인 상품은 월간 랭킹에서 제외 (1~2회 등장한 상품이 평균순위만으로 상위권에 오르는 것 방지)
    # 단, 필터 후 20개가 안 되면 기준을 2회→1회로 순차 완화해 최종 20개를 채운다
    min_appear = min(3, total_days)
    filtered = grouped[grouped["등장횟수"] >= min_appear]
    while len(filtered) < 20 and min_appear > 1:
        min_appear -= 1
        filtered = grouped[grouped["등장횟수"] >= min_appear]
    grouped = filtered

    # 월간점수 = 평균점수×0.7 + 등장률×5점×0.3  (순위 70% + 꾸준함 30%)
    grouped["평균점수"] = grouped["점수합"] / grouped["등장횟수"]
    grouped["등장률"]   = grouped["등장횟수"] / total_days
    grouped["월간점수"] = grouped["평균점수"] * 0.7 + (grouped["등장률"] * 5) * 0.3
    grouped = grouped.sort_values("월간점수", ascending=False).head(20).reset_index(drop=True)
    grouped.insert(0, "순위(월평균)", range(1, len(grouped) + 1))
    grouped["평균순위"] = grouped["평균순위"].round(2)
    grouped["평균가격"] = grouped["평균가격_raw"].apply(
        lambda x: f"{int(x):,}원" if x > 0 else "-"
    )
    grouped["카테고리"] = grouped.apply(
        lambda r: classify(r["상품명"], r["브랜드"]), axis=1
    )

    # 전월 대비 순위·가격 변동 (상품명+브랜드로 매칭, 전월 TOP20에 없으면 신규 진입 🆕)
    prev_lookup = {}
    if prev_df is not None and not prev_df.empty:
        for _, r in prev_df.iterrows():
            prev_lookup[(r["상품명"], r["브랜드"])] = (
                r["순위(월평균)"], parse_saved_price(r.get("평균가격"))
            )

    전월순위_list, 순위변동_list, 전월가격_list, 가격변동_list = [], [], [], []
    for _, r in grouped.iterrows():
        prev_rank, prev_price = prev_lookup.get((r["상품명"], r["브랜드"]), (None, None))
        전월순위_list.append(prev_rank)  # 없으면 None(NaN) — 숫자 컬럼에 "-" 문자열을 섞지 않기 위함
        순위변동_list.append(rank_change_label(prev_rank, r["순위(월평균)"]))
        전월가격_list.append(f"{int(prev_price):,}원" if prev_price else "-")
        가격변동_list.append(price_change_label(prev_price, r["평균가격_raw"]))

    grouped["전월순위"] = 전월순위_list
    grouped["순위변동"] = 순위변동_list
    grouped["전월가격"] = 전월가격_list
    grouped["가격변동"] = 가격변동_list

    return grouped[COLUMNS_OUT]


def build_monthly_category_stats(platform_dfs: dict, prev_stats_df: pd.DataFrame = None) -> pd.DataFrame:
    counts = defaultdict(lambda: {p: 0 for p in PLATFORM_NAMES})
    total  = 0

    for platform, df in platform_dfs.items():
        if df.empty:
            continue
        for _, row in df.iterrows():
            cat = classify(row.get("상품명", ""), row.get("브랜드", ""))
            counts[cat][platform] += 1
            total += 1

    prev_ratio = {}
    if prev_stats_df is not None and not prev_stats_df.empty:
        for _, r in prev_stats_df.iterrows():
            prev_ratio[r["카테고리"]] = r.get("비율(%)")

    rows = []
    for cat in ALL_CATEGORIES:
        c = counts.get(cat, {p: 0 for p in PLATFORM_NAMES})
        n = sum(c.values())
        ratio = round(n / total * 100, 1) if total else 0.0
        prev_r = prev_ratio.get(cat)
        rows.append({
            "카테고리":      cat,
            "전체":          n,
            "비율(%)":       ratio,
            "전월비율(%)":   prev_r,  # 전월 데이터 없으면 None(NaN)
            "증감(%p)":      round(ratio - prev_r, 1) if prev_r is not None else None,
            "카카오선물하기": c.get("카카오선물하기", 0),
            "다이소몰":       c.get("다이소몰", 0),
            "올리브영":       c.get("올리브영", 0),
        })

    rows.append({
        "카테고리":      "합계",
        "전체":          total,
        "비율(%)":       100.0,
        "전월비율(%)":   prev_ratio.get("합계"),
        "증감(%p)":      None,
        "카카오선물하기": sum(r["카카오선물하기"] for r in rows),
        "다이소몰":       sum(r["다이소몰"]       for r in rows),
        "올리브영":       sum(r["올리브영"]       for r in rows),
    })

    return pd.DataFrame(rows)


def build_monthly_brand_stats(platform_dfs: dict, prev_brand_df: pd.DataFrame = None) -> pd.DataFrame:
    """플랫폼별 TOP20 내 브랜드 등장 횟수를 집계 (브랜드 파워 추이 파악용)."""
    counts = defaultdict(lambda: {p: 0 for p in PLATFORMS})
    total = 0

    for platform, df in platform_dfs.items():
        if df.empty:
            continue
        for _, row in df.iterrows():
            brand = str(row.get("브랜드") or "").strip() or "미상"
            counts[brand][platform] += 1
            total += 1

    prev_counts = {}
    if prev_brand_df is not None and not prev_brand_df.empty:
        for _, r in prev_brand_df.iterrows():
            prev_counts[r["브랜드"]] = r.get("등장횟수(TOP20 합산)")

    rows = []
    for brand, c in counts.items():
        n = sum(c.values())
        prev_n = prev_counts.get(brand)
        rows.append({
            "브랜드":               brand,
            "등장횟수(TOP20 합산)": n,
            "비율(%)":              round(n / total * 100, 1) if total else 0.0,
            "전월등장횟수":         prev_n,  # 전월 데이터 없으면 None(NaN)
            "증감":                 (int(n - prev_n) if prev_n is not None else None),
            "카카오선물하기":       c.get("카카오선물하기", 0),
            "다이소몰":             c.get("다이소몰", 0),
            "올리브영":             c.get("올리브영", 0),
        })

    rows.sort(key=lambda r: r["등장횟수(TOP20 합산)"], reverse=True)
    return pd.DataFrame(rows)


def build_key_metrics(platform_dfs: dict, brand_stats_df: pd.DataFrame,
                       category_stats_df: pd.DataFrame, total_days: int) -> pd.DataFrame:
    """수집일수·플랫폼별 평균가격·전체 상품수/브랜드수/카테고리수 등 이번 달 숫자를 한눈에 보여준다."""
    rows = [{"항목": "수집일수", "값": f"{total_days}일"}]

    total_products = 0
    for platform in PLATFORMS:
        df = platform_dfs.get(platform, pd.DataFrame())
        total_products += len(df)
        prices = [parse_saved_price(p) for p in df.get("평균가격", [])] if not df.empty else []
        prices = [p for p in prices if p is not None]
        avg = round(sum(prices) / len(prices)) if prices else None
        rows.append({"항목": f"{platform} 평균가격(TOP20)", "값": f"{avg:,}원" if avg else "-"})

    rows.append({"항목": "전체 상품수(TOP20 합산)", "값": f"{total_products}개"})

    if brand_stats_df is not None and not brand_stats_df.empty:
        rows.append({"항목": "전체 브랜드수", "값": f"{len(brand_stats_df)}개"})

    if category_stats_df is not None and not category_stats_df.empty:
        n_cat = len(category_stats_df[
            (category_stats_df["카테고리"] != "합계") & (category_stats_df["전체"] > 0)
        ])
        rows.append({"항목": "등장 카테고리수", "값": f"{n_cat}개"})

    return pd.DataFrame(rows)


def build_platform_top1(platform_dfs: dict) -> pd.DataFrame:
    """플랫폼별 이번 달 평균순위 1위 상품."""
    rows = []
    for platform, df in platform_dfs.items():
        if df.empty:
            continue
        top = df[df["순위(월평균)"] == 1]
        if top.empty:
            continue
        r = top.iloc[0]
        rows.append({
            "플랫폼": platform, "상품명": r["상품명"], "브랜드": r["브랜드"],
            "카테고리": r["카테고리"], "평균가격": r["평균가격"],
        })
    return pd.DataFrame(rows)


def build_category_recommendations(platform_dfs: dict, category_stats_df: pd.DataFrame, top_n: int = 5):
    """비중이 늘고(성장) 평균순위도 좋은(고순위) 카테고리를 자동 선별해 대표 상품 예시와 함께 제안한다.
    '기타/특수목적'은 미분류 상품을 모아둔 카테고리라 추천 대상에서 제외한다.
    비교할 전월 데이터가 없으면 안내 문구를 대신 반환한다."""
    all_rows = []
    for platform, df in platform_dfs.items():
        if df.empty:
            continue
        tmp = df[df["카테고리"] != "기타/특수목적"][["카테고리", "상품명", "브랜드", "평균순위"]].copy()
        tmp["플랫폼"] = platform
        all_rows.append(tmp)
    if not all_rows or not any(len(t) for t in all_rows):
        return "데이터가 없어 카테고리 추천을 계산할 수 없습니다"

    combined = pd.concat(all_rows, ignore_index=True)
    cat_rank = combined.groupby("카테고리").agg(
        카테고리평균순위=("평균순위", "mean"),
        등장상품수=("상품명", "count"),
    ).reset_index()
    cat_rank["카테고리평균순위"] = cat_rank["카테고리평균순위"].round(2)

    if category_stats_df is None or category_stats_df.empty:
        return "전월 데이터가 없어 카테고리 추천(성장+고순위 기준)을 계산할 수 없습니다"

    growth = category_stats_df[
        (category_stats_df["카테고리"] != "합계") & category_stats_df["증감(%p)"].notna()
    ][["카테고리", "증감(%p)"]]
    merged = cat_rank.merge(growth, on="카테고리", how="inner")
    if merged.empty:
        return "전월 대비 성장 데이터가 없어 카테고리 추천을 계산할 수 없습니다"

    grown = merged[merged["증감(%p)"] > 0].sort_values(
        ["증감(%p)", "카테고리평균순위"], ascending=[False, True]
    ).head(top_n)
    if grown.empty:
        return "이번 달 비중이 늘어난 카테고리가 없어 추천할 항목이 없습니다"

    examples = []
    for cat in grown["카테고리"]:
        best = combined[combined["카테고리"] == cat].sort_values("평균순위").iloc[0]
        examples.append((best["상품명"], best["브랜드"], best["플랫폼"]))

    grown = grown.copy()
    grown["대표상품"], grown["대표브랜드"], grown["대표플랫폼"] = zip(*examples)
    grown = grown.rename(columns={"증감(%p)": "비중증감(%p)"})
    return grown[["카테고리", "비중증감(%p)", "카테고리평균순위", "등장상품수", "대표상품", "대표브랜드", "대표플랫폼"]]


def build_growth_insights(rising: pd.DataFrame, cat_up: pd.DataFrame, new_df: pd.DataFrame,
                           top_n: int = 5) -> pd.DataFrame:
    """급성장 카테고리·급상승 제품·신규 진입 상품을 신제품 기획 관점에서 한 줄로 재해석한다.
    새로운 지표를 계산하지 않고 이미 산출된 세 블록을 재구성만 한다.
    신규 진입은 전월 데이터가 없는 첫 달에 TOP20 전체가 한꺼번에 잡혀 표가 지나치게 길어지는 것을 막기 위해
    순위가 가장 좋은 top_n개만 노출한다(전체 목록은 위 '신규 진입 상품' 블록에서 확인)."""
    rows = []

    for _, r in cat_up.iterrows():
        rows.append({
            "구분": "급성장 카테고리",
            "항목": r["카테고리"],
            "핵심수치": f"{r['전월비율(%)']:.1f}%→{r['비율(%)']:.1f}% ({r['증감(%p)']:+.1f}%p)",
            "신제품 시사점": f"'{r['카테고리']}' 수요 확대 — 관련 원료·제형 신제품 검토 우선순위로",
        })

    for _, r in rising.iterrows():
        rows.append({
            "구분": "급성장 제품",
            "항목": f"[{r['플랫폼']}] {r['브랜드']} - {r['상품명']}",
            "핵심수치": f"{r['전월순위']}위 → {r['이번달순위']}위",
            "신제품 시사점": "가격·구성·마케팅 포인트 벤치마킹 대상",
        })

    for _, r in new_df.sort_values("이번달순위").head(top_n).iterrows():
        rows.append({
            "구분": "신규 진입(특이사항 후보)",
            "항목": f"[{r['플랫폼']}] {r['브랜드']} - {r['상품명']}",
            "핵심수치": f"이번 달 {r['이번달순위']}위 신규 진입",
            "신제품 시사점": "신제형·신규 구성 여부 확인 후 벤치마킹 대상 추가 검토",
        })

    return pd.DataFrame(rows, columns=["구분", "항목", "핵심수치", "신제품 시사점"])


def build_monthly_summary(platform_dfs: dict, category_stats_df: pd.DataFrame,
                           brand_stats_df: pd.DataFrame, total_days: int, top_n: int = 5) -> list:
    """핵심 지표, 플랫폼별 1위, 카테고리 추천, 급상승/급하락 TOP5, 카테고리 증감 TOP5, 신규 진입을 모은 요약 블록 목록."""
    change_rows = []
    new_rows = []
    for platform, df in platform_dfs.items():
        if df.empty:
            continue
        for _, r in df.iterrows():
            if r["순위변동"] == "🆕":
                new_rows.append({
                    "플랫폼": platform, "상품명": r["상품명"], "브랜드": r["브랜드"],
                    "이번달순위": r["순위(월평균)"],
                })
                continue
            change_rows.append({
                "플랫폼": platform, "상품명": r["상품명"], "브랜드": r["브랜드"],
                "전월순위": int(r["전월순위"]), "이번달순위": int(r["순위(월평균)"]),
                "순위변동폭": int(r["전월순위"]) - int(r["순위(월평균)"]),
            })

    change_cols = ["플랫폼", "상품명", "브랜드", "전월순위", "이번달순위"]
    change_df = pd.DataFrame(change_rows)
    if not change_df.empty:
        rising  = change_df.sort_values("순위변동폭", ascending=False).head(top_n)[change_cols]
        falling = change_df.sort_values("순위변동폭", ascending=True).head(top_n)[change_cols]
    else:
        rising = falling = pd.DataFrame(columns=change_cols)

    cat_cols = ["카테고리", "전월비율(%)", "비율(%)", "증감(%p)"]
    cat_rows = pd.DataFrame()
    if category_stats_df is not None and not category_stats_df.empty:
        cat_rows = category_stats_df[
            (category_stats_df["카테고리"] != "합계") & category_stats_df["증감(%p)"].notna()
        ]
    if not cat_rows.empty:
        cat_up   = cat_rows.sort_values("증감(%p)", ascending=False).head(top_n)[cat_cols]
        cat_down = cat_rows.sort_values("증감(%p)", ascending=True).head(top_n)[cat_cols]
    else:
        cat_up = cat_down = pd.DataFrame(columns=cat_cols)

    new_df = pd.DataFrame(new_rows, columns=["플랫폼", "상품명", "브랜드", "이번달순위"])
    insight_df = build_growth_insights(rising, cat_up, new_df, top_n)

    return [
        ("이번 달 핵심 지표", build_key_metrics(platform_dfs, brand_stats_df, category_stats_df, total_days)),
        ("플랫폼별 1위 상품", build_platform_top1(platform_dfs)),
        ("성장 + 고순위 카테고리 추천", build_category_recommendations(platform_dfs, category_stats_df, top_n)),
        (f"전월 대비 급상승 TOP{top_n}", rising),
        (f"전월 대비 급하락 TOP{top_n}", falling),
        (f"카테고리 비중 급증 TOP{top_n}", cat_up),
        (f"카테고리 비중 급감 TOP{top_n}", cat_down),
        (f"신규 진입 상품 ({len(new_df)}개)", new_df),
        ("신제품 기획 인사이트 (급성장·특이사항 재해석)", insight_df),
    ]


def write_sheet(writer, sheet_name: str, df: pd.DataFrame):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def write_summary_sheet(writer, sheet_name: str, blocks: list):
    """제목 + 표 여러 개를 한 시트에 세로로 이어붙인다 (전월 데이터 없으면 '해당 없음')."""
    from openpyxl.styles import Font

    ws = writer.book.create_sheet(sheet_name, 0)
    row = 1
    for title, df in blocks:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1
        if isinstance(df, str):
            ws.cell(row=row, column=1, value=df)
            row += 2
            continue
        if df is None or df.empty:
            ws.cell(row=row, column=1, value="해당 없음")
            row += 2
            continue
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col_idx, value=col_name).font = Font(bold=True)
        row += 1
        for _, r in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=row, column=col_idx, value=r[col_name])
            row += 1
        row += 1

    for col_cells in ws.columns:
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        if lengths:
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(lengths) + 4, 60)


def main():
    today = date.today()
    if today.month == 1:
        target_year, target_month = today.year - 1, 12
    else:
        target_year, target_month = today.year, today.month - 1

    prefix = f"{target_year}-{target_month:02d}"
    print(f"===== {prefix} 월별 취합 시작 =====")

    daily_dir  = Path("data/daily") / prefix
    daily_files = sorted(daily_dir.glob(f"{prefix}-*.xlsx")) if daily_dir.exists() else []

    if not daily_files:
        print(f"  데이터 없음: data/daily/{prefix}/ 에 파일이 없습니다")
        return

    total_days = len(daily_files)  # 실제 수집된 일수 기준
    print(f"  일별 파일 {total_days}개 발견")

    prev_sheets = load_previous_month(target_year, target_month)
    if prev_sheets:
        print("  전월 취합본 발견 → 순위/가격/카테고리/브랜드 전월 대비 비교 반영")

    output_path = Path("data/monthly") / f"{prefix}_월별취합.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    platform_dfs = {
        platform: aggregate_platform(daily_files, platform, total_days, prev_sheets.get(platform))
        for platform in PLATFORMS
    }
    stats_df = build_monthly_category_stats(platform_dfs, prev_sheets.get("카테고리통계"))
    brand_df = build_monthly_brand_stats(platform_dfs, prev_sheets.get("브랜드통계"))
    summary_blocks = build_monthly_summary(platform_dfs, stats_df, brand_df, total_days)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_summary_sheet(writer, "월별요약", summary_blocks)
        for platform in PLATFORMS:
            write_sheet(writer, platform, platform_dfs[platform])
        write_sheet(writer, "카테고리통계", stats_df)
        write_sheet(writer, "브랜드통계", brand_df)

    print(f"저장 완료: {output_path}")


if __name__ == "__main__":
    main()
