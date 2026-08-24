"""엑셀에 제품 이미지를 넣는 공통 유틸 — 두 가지 보기를 제공한다.

1. `insert_image_column()` — 기존 표(세로 1행=1상품) 시트 A열에 작은 썸네일을 끼워넣음.
   월별집계·health-trend가 그대로 읽는 원본 데이터 시트라 컬럼 구성을 유지해야 함.
2. `insert_card_sheet()` — 순위를 가로로 펼쳐 열마다 상품 하나가 되는 별도 '카드형' 시트를
   새로 추가. 사람이 훑어보기 좋도록 이미지를 훨씬 크게(기본 130px) 넣는 용도이며, 원본
   표 시트는 건드리지 않으므로 월별집계·health-trend 로직에는 영향 없음.

이미지 자체는 각 크롤러가 이미 수집해둔 이미지URL 컬럼을 재사용한다(재수집 없음). 여기서는
그 URL을 다운로드해 openpyxl 이미지로 셀에 임베드하는 역할만 한다(health-trend
대시보드처럼 <img src=URL>로 실시간 참조하는 방식은 엑셀에서는 불가능해 다운로드 임베드가
유일한 방법).
"""

import io

import requests
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from crawlers.base import USER_AGENT

# --- 표 시트 A열 썸네일 (작게, 데이터 행 높이에 맞춤) ---
THUMB_PX = 48
THUMB_SOURCE_PX = 96
ROW_HEIGHT_PT = 38

# --- 카드형 시트 (크게, 이미지 자체가 주인공) ---
CARD_IMG_PX = 130
CARD_IMG_SOURCE_PX = 280
CARD_ROW_HEIGHT_PT = 100
CARD_NAME_ROW_HEIGHT_PT = 46
CARD_COL_WIDTH = 17
CARD_LABEL_COL_WIDTH = 10

# 원본 상품 이미지(수백 KB~수 MB급)를 셀 표시 크기로만 리사이즈하지 않고 그대로 임베드하면
# 일별 xlsx 하나가 수십 MB로 불어난다(실측: 48px 표시로 40장 임베드 시 55MB). 화면엔 작게만
# 보이므로 다운로드 단계에서부터 리샘플링 + JPEG 재압축해서 넣는다.
JPEG_QUALITY = 80


def _fetch_raw(url: str, cache: dict = None):
    """원본 이미지 바이트를 받아온다. cache(dict)를 넘기면 같은 URL 재요청을 건너뛴다 —
    표 시트(작게)와 카드형 시트(크게) 둘 다 같은 원본을 두 번 내려받게 되는데, 백필처럼
    한 프로세스에서 대량으로 돌릴 때 이 캐시 하나로 다운로드 횟수를 절반으로 줄인다."""
    if cache is not None and url in cache:
        return cache[url]
    try:
        resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=10)
        resp.raise_for_status()
        data = resp.content
    except Exception:
        data = None
    if cache is not None:
        cache[url] = data
    return data


def _download_thumbnail(url: str, max_px: int = THUMB_SOURCE_PX, cache: dict = None):
    """URL을 받아 max_px 이내로 축소·재압축된 JPEG 버퍼와 (width, height)를 반환. 실패 시 None."""
    if not url or not isinstance(url, str):
        return None
    raw = _fetch_raw(url, cache=cache)
    if raw is None:
        return None
    try:
        pil_img = PILImage.open(io.BytesIO(raw))
        pil_img = pil_img.convert("RGB")
        pil_img.thumbnail((max_px, max_px))
        buf = io.BytesIO()
        pil_img.save(buf, format="JPEG", quality=JPEG_QUALITY)
        buf.seek(0)
        return buf, pil_img.size
    except Exception:
        return None


def _add_scaled_image(ws, url: str, anchor: str, display_px: int, source_px: int, cache: dict = None) -> None:
    """다운로드 + 리샘플링 후 anchor 셀(예: "C2")에 display_px 이내로 맞춰 삽입. 실패 시 조용히 스킵."""
    result = _download_thumbnail(url, max_px=source_px, cache=cache)
    if result is None:
        return
    buf, (w, h) = result
    try:
        img = XLImage(buf)
        scale = display_px / max(w, h)
        img.width = round(w * scale)
        img.height = round(h * scale)
        ws.add_image(img, anchor)
    except Exception:
        pass


def hide_sheet(ws) -> None:
    """원본 표 시트를 탭에서 숨긴다(데이터는 그대로 남아 월별집계·health-trend가
    변함없이 읽을 수 있음 — pd.read_excel/openpyxl 모두 숨김 시트도 그대로 읽힘).
    사용자가 열었을 때 카드형·카테고리통계 탭만 보이게 하려고 2026-08-24 추가.
    엑셀에서 시트 탭 우클릭 → 숨기기 취소로 언제든 다시 보이게 할 수 있다."""
    ws.sheet_state = "hidden"


def insert_image_column(ws, image_urls: list, data_start_row: int = 2, header: str = "이미지", cache: dict = None) -> None:
    """워크시트 맨 앞(A열)에 이미지 컬럼을 새로 끼워넣고 image_urls[i]를
    (data_start_row + i) 행에 임베드한다. 개별 URL 다운로드 실패는 그 칸만 비우고
    조용히 넘어간다(한 상품 이미지 실패로 전체 저장이 막히지 않도록)."""
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value=header)
    ws.column_dimensions["A"].width = 9

    for i, url in enumerate(image_urls):
        row = data_start_row + i
        ws.row_dimensions[row].height = ROW_HEIGHT_PT
        _add_scaled_image(ws, url, f"A{row}", THUMB_PX, THUMB_SOURCE_PX, cache=cache)


_CARD_FIELD_LABELS = ["썸네일", "카테고리", "브랜드", "상품명", "판매가", "링크"]


def insert_card_sheet(wb, sheet_name: str, blocks: list, price_label: str = "판매가", cache: dict = None) -> None:
    """순위를 가로로 펼쳐 열마다 상품 하나가 되는 카드형 시트를 새로 만든다.
    blocks는 [(블록 라벨, items), ...] — 서브카테고리처럼 순위가 1위부터 여러 번 반복되는
    플랫폼(카카오선물하기)은 블록을 나눠 세로로 쌓아 1위~N위 헤더를 블록마다 따로 갖는다
    (한 줄로 펼치면 가로 스크롤이 너무 길어짐). 블록이 하나뿐이면 기존 단일 카드형과 동일.
    items는 카테고리/순위/상품명/브랜드/가격/상품URL/이미지URL 키를 가진 dict 리스트를 받는다
    (일별은 main.py의 platform_data 값 그대로, 월별취합은 순위(월평균)→순위·평균가격→가격으로
    맞춰 변환해서 넘김). price_label은 "가격" 행에 붙는 라벨만 바꾼다(월별취합은 "평균가격").
    기존 표 시트는 건드리지 않는다."""
    ws = wb.create_sheet(title=sheet_name)

    max_cols = max((len(items) for _, items in blocks), default=0)
    ws.column_dimensions["A"].width = CARD_LABEL_COL_WIDTH
    ws.column_dimensions["B"].width = CARD_LABEL_COL_WIDTH
    for i in range(max_cols):
        ws.column_dimensions[get_column_letter(3 + i)].width = CARD_COL_WIDTH

    row = 1
    for block_label, items in blocks:
        row = _write_card_block(ws, row, block_label, items, price_label, cache=cache)

    ws.freeze_panes = "C2"


def _write_card_block(ws, header_row: int, block_label: str, items: list, price_label: str, cache: dict = None) -> int:
    """header_row부터 카드 블록 하나(헤더+썸네일+필드 6줄)를 그리고, 다음 블록이 시작할
    행 번호(빈 줄 1개 포함)를 반환한다."""
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    field_labels = list(_CARD_FIELD_LABELS)
    field_labels[4] = price_label

    ws.cell(row=header_row, column=1, value="플랫폼").font = bold
    ws.cell(row=header_row, column=2, value="순위").font = bold
    for i, item in enumerate(items):
        ws.cell(row=header_row, column=3 + i, value=f"{item.get('순위', i + 1)}위").font = bold

    for r, label in enumerate(field_labels, start=header_row + 1):
        ws.cell(row=r, column=2, value=label).font = bold

    last_row = header_row + len(_CARD_FIELD_LABELS)
    thumb_row, name_row = header_row + 1, header_row + 4

    ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=last_row, end_column=1)
    plat_cell = ws.cell(row=header_row + 1, column=1, value=block_label)
    plat_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    plat_cell.font = Font(bold=True, size=12)

    ws.row_dimensions[thumb_row].height = CARD_ROW_HEIGHT_PT
    ws.row_dimensions[name_row].height = CARD_NAME_ROW_HEIGHT_PT

    for i, item in enumerate(items):
        col = 3 + i
        col_letter = get_column_letter(col)

        ws.cell(row=header_row + 2, column=col, value=item.get("카테고리", "")).alignment = center
        ws.cell(row=header_row + 3, column=col, value=item.get("브랜드", "")).alignment = center
        name_cell = ws.cell(row=name_row, column=col, value=item.get("상품명", ""))
        name_cell.alignment = center
        ws.cell(row=header_row + 5, column=col, value=item.get("가격", "")).alignment = center

        link_cell = ws.cell(row=header_row + 6, column=col)
        url = item.get("상품URL", "")
        if url:
            link_cell.value = "바로가기"
            link_cell.hyperlink = url
            link_cell.font = Font(color="0563C1", underline="single")
        link_cell.alignment = center

        _add_scaled_image(ws, item.get("이미지URL", ""), f"{col_letter}{thumb_row}", CARD_IMG_PX, CARD_IMG_SOURCE_PX, cache=cache)

    return last_row + 2  # 다음 블록 전 빈 줄 1개
